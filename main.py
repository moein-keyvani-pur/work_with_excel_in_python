import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def save_file(workbook, file_name):
    workbook.save(file_name)


def draw_chart_all_income(sheet):
    values = Reference(worksheet=sheet, min_row=2, max_row=sheet.max_row, min_col=6, max_col=6)
    chart = BarChart()
    chart.add_data(data=values)
    sheet.add_chart(chart, 'k4')


def calculate_all_income(sheet):
    for i in range(2, sheet.max_row + 1):
        hour = sheet.cell(i, 4).value
        income_hour = sheet.cell(i, 5).value
        all_income = hour * income_hour
        # print(f'all_income is {all_income}')
        sheet.cell(i, 6).value = all_income
        draw_chart_all_income(sheet)


def load_file(file_name):
    wb = xl.load_workbook(file_name)
    sheet = wb['Sheet1']
    calculate_all_income(sheet)
    save_file(workbook=wb, file_name=file_name)


file_name = 'info_employee.xlsx'
load_file(file_name=file_name)
